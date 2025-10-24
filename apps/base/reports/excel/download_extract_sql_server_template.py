from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook import Workbook
from apps.base.helpers.format_response import FormatResponse
import os

def download_extract_sql_server_template(all_sql_results):
    """
    Genera y descarga una plantilla Excel con los resultados de los archivos .rpt
    usando solo OpenPyXL (compatible con Python 32 bits).
    """
    try:
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="reporte-rpt.xlsx"'

        # Crear Workbook y seleccionar hoja activa
        wb = Workbook()
        ws = wb.active
        ws.title = 'report'

        # Columnas de la plantilla
        columns = ['Nombre del archivo', 'Ruta', 'Tipo', 'Descripción', 'Existe']

        # Encabezado azul oscuro con letra blanca
        header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF00008B', end_color='FF00008B', fill_type='solid')
        for idx, col in enumerate(columns):
            cell = ws.cell(row=1, column=idx+1, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[chr(65 + idx)].width = len(col) + 5

        # Escribir los datos directamente desde all_sql_results
        row_idx = 2
        for rpt_file, results in all_sql_results.items():
            for item in results:
                ws.cell(row=row_idx, column=1, value=item.get('file_name'))
                ws.cell(row=row_idx, column=2, value=item.get('file_route'))
                ws.cell(row=row_idx, column=3, value=item.get('type'))
                ws.cell(row=row_idx, column=4, value=item.get('descripcion_query'))
                ws.cell(row=row_idx, column=5, value=item.get('exist'))
                row_idx += 1

        # Guardar y enviar
        wb.save(response)
        return response

    except Exception as e:
        return FormatResponse.failed(e)
