from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook import Workbook


def download_jobs_excel(jobs_data):
    try:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="jobs-sql-report.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Jobs'

        #  Columnas nuevas basadas en tu modelo
        columns = [
            'Id Job',
            'Nombre Job',
            # 'Numero Paso',
            # 'Nombre Paso',
            'Accion',
            'Tabla',
            # 'Sentencia',
            'Base de Datos',
            'Stored Procedure'
        ]

        #  Header bonito
        header_font = Font(bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF1F4E78', end_color='FF1F4E78', fill_type='solid')

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

            ws.column_dimensions[chr(64 + col_idx)].width = 25

        #  APLANAR DATA (CLAVE)
        row_idx = 2

        for job in jobs_data:
            for step in job.get("Steps", []):
                detalles = step.get("DML_Detalle", [])

                #  si no hay detalle igual crea fila
                if not detalles:
                    ws.cell(row=row_idx, column=1, value=job.get("IdJob"))
                    ws.cell(row=row_idx, column=2, value=job.get("NombreJob"))
                    # ws.cell(row=row_idx, column=3, value=step.get("NumeroPaso"))
                    # ws.cell(row=row_idx, column=4, value=step.get("NombrePaso"))
                    ws.cell(row=row_idx, column=3, value=step.get("BaseDatos"))
                    ws.cell(row=row_idx, column=4, value=step.get("StoredProcedure"))
                    row_idx += 1
                    continue

                #  UNA FILA POR CADA DML_Detalle
                for d in detalles:
                    ws.cell(row=row_idx, column=1, value=job.get("IdJob"))
                    ws.cell(row=row_idx, column=2, value=job.get("NombreJob"))
                    # ws.cell(row=row_idx, column=3, value=step.get("NumeroPaso"))
                    # ws.cell(row=row_idx, column=4, value=step.get("NombrePaso"))
                    ws.cell(row=row_idx, column=3, value=d.get("action"))
                    ws.cell(row=row_idx, column=4, value=d.get("table"))
                    # ws.cell(row=row_idx, column=5, value=d.get("statement"))
                    ws.cell(row=row_idx, column=5, value=step.get("BaseDatos"))
                    ws.cell(row=row_idx, column=6, value=step.get("StoredProcedure"))

                    row_idx += 1

        wb.save(response)
        return response

    except Exception as e:
        return HttpResponse(str(e), status=500)